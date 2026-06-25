"""
Recibe JSON por stdin con { proforma_id, items[], template_path }
Llena la hoja 1 del template con los items y devuelve el .xlsx por stdout binario.
"""
import sys
import json
import io
import os
import openpyxl
from openpyxl.cell.cell import MergedCell
from copy import copy

def copy_cell_style(src_cell, dst_cell):
    """Copia formato de celda fuente a destino."""
    if src_cell.has_style:
        dst_cell.font      = copy(src_cell.font)
        dst_cell.fill      = copy(src_cell.fill)
        dst_cell.border    = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

def main():
    data = json.loads(sys.stdin.read())
    items         = data.get("items", [])
    template_path = data.get("template_path", "")
    proforma_id   = data.get("proforma_id", "")
    cliente_nombre = data.get("cliente_nombre", "CLIENTE")
    asesor_nombre  = data.get("asesor_nombre", "ASESOR")
    tipo           = data.get("tipo", "MARITIMA").upper()
    fecha          = data.get("fecha", "")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    # Fila plantilla a copiar formato (fila 3)
    TEMPLATE_ROW = 3
    DATA_START   = 3   # primera fila de datos
    MAX_COLS     = 31  # columnas A-AE

    # Guardar estilos de la fila plantilla
    template_styles = {}
    for col in range(1, MAX_COLS + 1):
        cell = ws.cell(row=TEMPLATE_ROW, column=col)
        if cell.has_style:
            template_styles[col] = {
                "font":      copy(cell.font),
                "fill":      copy(cell.fill),
                "border":    copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
            }

    # Limpiar filas de datos existentes
    for r in range(DATA_START, DATA_START + 50):
        for col in range(1, MAX_COLS + 1):
            cell = ws.cell(row=r, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # Escribir items
    for i, item in enumerate(items):
        r = DATA_START + i

        # Aplicar estilo de fila plantilla
        for col in range(1, MAX_COLS + 1):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, MergedCell):
                continue
            s = template_styles.get(col)
            if s:
                cell.font      = copy(s["font"])
                cell.fill      = copy(s["fill"])
                cell.border    = copy(s["border"])
                cell.alignment = copy(s["alignment"])
                cell.number_format = s["number_format"]

        imagen_link = ""
        if item.get("imagen_drive_id"):
            imagen_link = f"https://drive.google.com/uc?export=view&id={item['imagen_drive_id']}"

        def set_val(col, value):
            cell = ws.cell(row=r, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = value

        # Mapeo de columnas
        set_val(1,  item.get("link_cotizador", ""))          # A LINK COTIZADOR
        set_val(2,  1)                                                               # B PROVEEDORES
        set_val(3,  "")                                                             # C PROVEEDORES 2
        set_val(4,  item.get("modelo") or "")                                      # D MODELO
        set_val(5,  "")                                                             # E FOTO (vacío)
        set_val(6,  imagen_link)                                                    # F LINK IMAGEN
        set_val(7,  item.get("descripcion") or item.get("nombre_comercial") or "") # G DESCRIPCION
        set_val(8,  item.get("nombre_comercial") or "")                            # H NOMBRE COMERCIAL
        set_val(9,  item.get("unidad_medida") or "PZA")                           # I UNIDAD DE MEDIDA
        set_val(10, item.get("cantidad_x_caja") or 1)                             # J CANTIDAD x CAJA
        set_val(11, item.get("cajas") or 1)                                       # K CAJAS
        set_val(12, item.get("total_unidades"))                                    # L TOTAL UNIDADES
        set_val(13, None)                                                          # M PRECIO YUANES
        set_val(14, item.get("valor_unitario_usd"))                               # N PRECIO UNIT USD
        set_val(15, item.get("valor_total_usd"))                                  # O TOTAL USD
        # P-T: CBM, peso, flete — dejar vacío
        set_val(21, item.get("partida") or "")                                    # U PARTIDA
        set_val(22, item.get("tnan") or "0000")                                   # V TNAN
        set_val(23, item.get("arancel", 0.3))                                     # W ARANCEL %
        set_val(24, item.get("arancel_tlc", 0.3))                                 # X ARANCEL TLC %
        set_val(25, "NO")                                                          # Y PERMISOS
        set_val(30, "NO")                                                          # AD PERMISOS 2
        set_val(31, item.get("nombre_comercial") or "")                            # AE NOMBRE COMERCIAL

    # Ajustar nombre de hoja y archivo
    ws.title = "1.CÓL"

    output = io.BytesIO()
    wb.save(output)
    sys.stdout.buffer.write(output.getvalue())

if __name__ == "__main__":
    main()
